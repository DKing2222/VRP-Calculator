import tkinter
from math import *
from tkinter import filedialog, ttk

import customtkinter
import tkintermapview
from tkintermapview import TkinterMapView
import random
import pandas
from openpyxl import Workbook
from openpyxl import load_workbook

from pyroutelib3 import Router

#from src.pyVRP import  build_coordinates, build_distance_matrix, genetic_algorithm_vrp, plot_tour_coordinates
#from verypy.classic_heuristics.parallel_savings import parallel_savings_init
#from verypy.util import sol2routes



customtkinter.set_default_color_theme("blue")
customtkinter.set_appearance_mode('Dark')
random.seed(1)


class App(customtkinter.CTk,tkinter.Tk):

    APP_NAME = "DK VRP Solver"

    WIDTH = 800

    HEIGHT = 500



    def __init__(self, *args, **kwargs):

        super().__init__(*args, **kwargs)



        self.title(App.APP_NAME)

        self.geometry(str(App.WIDTH) + "x" + str(App.HEIGHT))

        self.minsize(App.WIDTH, App.HEIGHT)



        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.bind("<Command-q>", self.on_closing)

        self.bind("<Command-w>", self.on_closing)

        self.createcommand('tk::mac::Quit', self.on_closing)



        self.marker_list = []
        #self.depot=[]
        self.sommet_list =[]
        self.sommet_listC =[]
        self.matrice_distance=[]
        self.customer=[]
        self.matrice_savings=[]
        self.routeC=[]
        self.route=[]
        self.polygone= []
        self.cluster= []

        self.wb = Workbook() 




        # ============ create two CTkFrames ============



        self.grid_columnconfigure(0, weight=0)

        self.grid_columnconfigure(1, weight=1)

        self.grid_rowconfigure(0, weight=1)



        self.frame_left = customtkinter.CTkFrame(master=self, width=150, corner_radius=0, fg_color=None)

        self.frame_left.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")



        self.frame_right = customtkinter.CTkFrame(master=self, corner_radius=0)

        self.frame_right.grid(row=0, column=1, rowspan=1, pady=0, padx=0, sticky="nsew")



        # ============ frame_left ============



        self.frame_left.grid_rowconfigure(2, weight=1)



        self.button_1 = customtkinter.CTkButton(master=self.frame_left,

                                                text="Configurer Depot",

                                                command=self.depot_toplevel)

        self.button_1.grid(pady=(20, 0), padx=(20, 20), row=0, column=0)

        self.button_1 = customtkinter.CTkButton(master=self.frame_left,

                                                text="Ajouter Client",

                                                command=self.client_toplevel)

        self.button_1.grid(pady=(20, 0), padx=(20, 20), row=1, column=0)

        self.button_1 = customtkinter.CTkButton(master=self.frame_left,

                                                text="Liste des clients",

                                                command=self.tableau_clients)

        self.button_1.grid(pady=(20, 0), padx=(20, 20), row=2, column=0, sticky="n")

        self.button_2 = customtkinter.CTkButton(master=self.frame_left,

                                                text="Nettoyer",

                                                command=self.clear_marker_event)
        #self.button_2.configure(sticky)
        self.button_2.grid(pady=(20, 0), padx=(20, 20), row=3, column=0, sticky="n")



        self.adresse_label = customtkinter.CTkLabel(self.frame_left, text="", anchor="n")

        self.adresse_label.grid(row=4, column=0, padx=(20, 20), pady=(20, 0))

        self.map_label = customtkinter.CTkLabel(self.frame_left, text="Tile Server:", anchor="w")

        self.map_label.grid(row=5, column=0, padx=(20, 20), pady=(20, 0))

        self.map_option_menu = customtkinter.CTkOptionMenu(self.frame_left, values=["OpenStreetMap", "Google normal", "Google satellite"],

                                                                       command=self.change_map)

        self.map_option_menu.grid(row=6, column=0, padx=(20, 20), pady=(10, 0))



        self.appearance_mode_label = customtkinter.CTkLabel(self.frame_left, text="Apparence :", anchor="w")

        self.appearance_mode_label.grid(row=7, column=0, padx=(20, 20), pady=(20, 0))

        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.frame_left, values=["Light", "Dark", "System"],

                                                                       command=self.change_appearance_mode)

        self.appearance_mode_optionemenu.grid(row=8, column=0, padx=(20, 20), pady=(10, 20))



        # ============ frame_right ============



        self.frame_right.grid_rowconfigure(1, weight=1)

        self.frame_right.grid_rowconfigure(0, weight=0)

        self.frame_right.grid_columnconfigure(0, weight=1)

        self.frame_right.grid_columnconfigure(1, weight=0)

        self.frame_right.grid_columnconfigure(2, weight=1)
        self.frame_right.grid_columnconfigure(3, weight=0)
        self.frame_right.grid_columnconfigure(4, weight=0)



        self.map_widget = TkinterMapView(self.frame_right, corner_radius=0)
        #self.map_widget.set_position(-1.6586697, 29.1826432)

        self.map_widget.grid(row=1, rowspan=1, column=0, columnspan=5, sticky="nswe", padx=(0, 0), pady=(0, 0))

        def left_click_event(coordinates_tuple):
            print("Left click event with coordinates:", coordinates_tuple)
            self.adresse_label.set_text(self.adresse(coordinates_tuple[0],coordinates_tuple[1]))
            self.adresse_label.after(2000, self.adresse_label.set_text,"LON= "+str(round(coordinates_tuple[0],11)))
            self.adresse_label.after(4000, self.adresse_label.set_text,"LAT = "+str(round(coordinates_tuple[1],11)))
            self.adresse_label.after(6000, self.adresse_label.set_text,"")

    
        #self.map_widget.add_left_click_map_command(left_click_event)

        self.map_widget.add_right_click_menu_command(label="Configurer depot ici",
                                        command=self.auto_depot_toplevel,
                                        pass_coords=True)
        
        self.map_widget.add_right_click_menu_command(label="Ajouter client ici",
                                        command=self.auto_client_toplevel,
                                        pass_coords=True)


        self.entry = customtkinter.CTkEntry(master=self.frame_right,

                                            placeholder_text="Tapez l'adresse")

        self.entry.grid(row=0, column=0, sticky="we", padx=(12, 0), pady=12)

        #DK#self.entry.entry.bind("<Return>", self.search_event)



        self.button_5 = customtkinter.CTkButton(master=self.frame_right,

                                                text="Recherche",

                                                width=90,

                                                command=self.search_event)

        self.button_5.grid(row=0, column=1, sticky="w", padx=(12, 0), pady=12)

        self.button_6 = customtkinter.CTkButton(master=self.frame_right,

                                                text="Se positionner au depot",

                                                width=90,

                                                command=self.position_on_depot,

                                                )

        self.button_6.grid(row=0, column=2, sticky="e", padx=(12, 24), pady=12)


        self.button_6 = customtkinter.CTkButton(master=self.frame_right,

                                                text="Calculer",

                                                width=90,

                                                command=self.dessiner_route,
                                                
                                                hover_color="#222",

                                                )

        self.button_6.grid(row=0, column=3, sticky="e", padx=(12, 24), pady=12)




        # Set default values

        self.map_widget.set_address("Goma")

        self.map_option_menu.set("OpenStreetMap")

        self.appearance_mode_optionemenu.set("Dark")



    def search_event(self, event=None):

        self.map_widget.set_address(self.entry.get())



    def set_marker_event(self):

        #current_position = self.map_widget.get_position()
        if not self.sommet_list:
            self.point_dec=[float(self.long.get()),float(self.lat.get()), float(self.camion.get()),"Depot"]
            
            self.current_position = self.point_dec

            self.marker_list.append(self.map_widget.set_marker(self.current_position[0], self.current_position[1], text=self.current_position[3], marker_color_outside= "#ccf"))

            self.window.destroy()

            
            #self.sommet_list.pop(0)
            self.sommet_list.insert(0,self.point_dec)
        
        else:
            tkinter.messagebox.showerror(title='Erreur', message="Le depot est deja configuré, Nottoyer pour reconfigurer")
        #print(self.sommet_list)

    def auto_set_marker_event(self):

        #current_position = self.map_widget.get_position()

        if not self.sommet_list :

            self.point_dec=[self.longit,self.latit, float(self.camion.get()),"Depot"]
        
            self.current_position = self.point_dec

            self.marker_list.append(self.map_widget.set_marker(self.current_position[0], 
                                    self.current_position[1], text=self.current_position[3],
                                    marker_color_outside= "#007acc",marker_color_circle="#eef",
                                    text_color="#222"))


            if self.sommet_list:
                self.sommet_list.pop(0)
            self.sommet_list.insert(0,self.point_dec)
            #print(self.sommet_list)
        else:
            tkinter.messagebox.showerror(title='Erreur', message="Le depot est deja configuré, Nottoyer pour reconfigurer")
            #print(self.sommet_list)

        self.window.destroy()

    def set_marker_eventC(self):

        if self.sommet_list:

            self.point_decC=[float(self.longi.get()),float(self.lati.get()), float(self.demande.get()), str(self.nom.get())]
            current_positionC = self.point_decC
            self.sommet_listC.append(self.point_decC)

            self.marker_list.append(self.map_widget.set_marker(current_positionC[0], current_positionC[1], text=current_positionC[3]))
            
            # if self.sommet_listC[1]:
            #     self.dessiner_route()
        
        else:
            tkinter.messagebox.showerror(title='Erreur', message="Le depot n'est pas encore configuré; Configurer le depot")

        #path_1 = self.map_widget.set_path([(-1.6881886, 29.2368100), (-1.6440902, 29.2436764)])

        #print(self.sommet_list)
        self.windowC.destroy()

    def auto_set_marker_eventC(self):

        if self.sommet_list:

            self.point_decC=[self.longiti,self.latiti, float(self.demande.get()), str(self.nom.get())]
            
            if float(self.demande.get()) < self.sommet_list[0][2] :
                current_positionC = self.point_decC
                self.sommet_listC.append(self.point_decC)

                self.marker_list.append(self.map_widget.set_marker(current_positionC[0], current_positionC[1], text=current_positionC[3]))
                self.sommet_list.append(self.point_decC)
                self.windowC.destroy()
            else:
                tkinter.messagebox.showerror(title='Surcommande', message="Le client demmande plus que la capacite totale du camion, Veiller reserver une livraison special pour lui")
        
        else:
            tkinter.messagebox.showerror(title='Erreur', message="Le depot n'est pas encore configuré; Configurer le depot")
            self.windowC.destroy()

        #path_1 = self.map_widget.set_path([(-1.6881886, 29.2368100), (-1.6440902, 29.2436764)])

        #print(self.sommet_list)
        #self.windowC.destroy()


    def clear_marker_event(self):
        
        #self.sommet_list.clear()
        #print(self.sommet_list)

        for marker in self.marker_list:

            marker.delete()
        self.sommet_list.clear()
        self.sommet_listC.clear()
        self.clear_path()
        #print(self.sommet_list)
    

    def clear_path(self):
        
        #self.sommet_list.clear()
        #print(self.sommet_list)

        for path in self.polygone :

            path.delete()
        self.polygone.clear()
        #print(self.sommet_list)



    def change_appearance_mode(self, new_appearance_mode: str):

        customtkinter.set_appearance_mode(new_appearance_mode)



    def change_map(self, new_map: str):

        if new_map == "OpenStreetMap":

            self.map_widget.set_tile_server("https://a.tile.openstreetmap.org/{z}/{x}/{y}.png")

        elif new_map == "Google normal":

            self.map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)

        elif new_map == "Google satellite":

            self.map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)

#-----------------------DEPOT TOPLEVEL----------------------------

    def depot_toplevel(self):
        if not self.sommet_list:
            self.window = customtkinter.CTkToplevel(self)
            self.window.geometry("200x200")
            self.window.title("Coordonnées du Depot")
            #self.window.resizable(False,False)

            # self.titre_depot = customtkinter.CTkLabel(self.window, text="Depot & Camion:")

            # self.titre_depot.grid(row=0, column=0, sticky='e')

            # entrer la longitude
            self.long = customtkinter.CTkEntry(master=self.window,

                                                placeholder_text="Longitude")

            self.long.grid(row=1, column=1, sticky="we", padx=30, pady=6)

            #DK#self.long.entry.bind("<Return>", self.search_event)

            # entrer la latittude
            self.lat = customtkinter.CTkEntry(master=self.window,

                                                placeholder_text="Latittude")

            self.lat.grid(row=2, column=1, sticky="we", padx=30, pady=6)

            #DK#self.lat.entry.bind("<Return>", self.search_event)

            # entrer la Capacite du camion
            self.camion = customtkinter.CTkEntry(master=self.window,

                                                placeholder_text="Capacité du camion")

            self.camion.grid(row=3, column=1, sticky="we", padx=30, pady=6)

            #DK#self.camion.entry.bind("<Return>", self.set_marker_event)

            # Boutton valider

            self.valider_depot = customtkinter.CTkButton(master=self.window, text="Valider")

            self.valider_depot.configure(command=self.set_marker_event)

            self.valider_depot.grid(row=4, column=1, sticky="we", padx=30, pady=6)

        else:
            tkinter.messagebox.showerror(title='Erreur', message="Le depot est deja configuré, Nottoyer pour reconfigurer")
#--------------------------------------------------------------------------------

#-----------------------AUTO DEPOT TOPLEVEL----------------------------

    def auto_depot_toplevel(self,coordinates_tuple):
        if not self.sommet_list:

            self.window = customtkinter.CTkToplevel(self)
            self.window.geometry("200x200")
            self.window.title("Coordonnées du Depot")
            #self.window.resizable(0,0)

            # entrer la longitude
            self.long_label= customtkinter.CTkLabel(master= self.window, text="longitude: "+str(coordinates_tuple[0]))
            self.long_label.grid(row=0, column=1)
            self.longit=coordinates_tuple[0]

            # entrer la latittude
            self.lat_label= customtkinter.CTkLabel(master= self.window, text="longitude: "+str(coordinates_tuple[1]))
            self.lat_label.grid(row=1, column=1)
            self.latit=coordinates_tuple[1]

            # entrer la Capacite du camion
            self.camion = customtkinter.CTkEntry(master=self.window,

                                                placeholder_text="Capacité du camion")

            self.camion.grid(row=3, column=1, sticky="we", padx=30, pady=6)

            #DK#self.camion.entry.bind("<Return>", self.set_marker_event)

            # Boutton valider

            self.valider_depot = customtkinter.CTkButton(master=self.window, text="Valider")

            self.valider_depot.configure(command=self.auto_set_marker_event)

            self.valider_depot.grid(row=4, column=1, sticky="we", padx=30, pady=6)

        else:
            tkinter.messagebox.showerror(title='Erreur', message="Le depot est deja configuré, Nottoyer pour reconfigurer")
      
#--------------------------------------------------------------------------------

#-----------------------CLIENT TOPLEVEL----------------------------

    def client_toplevel(self):
        
        if not self.sommet_list:
            # self.dialog = customtkinter.CTkInputDialog(master=None, text="Type in a number:")
            tkinter.messagebox.showerror(title='Erreur', message="Le depot n'est pas encore configuré; Configurer le depot")
            self.depot_toplevel()
        else:
            self.windowC = customtkinter.CTkToplevel(self)
            self.windowC.geometry("200x220")
            self.windowC.title("Coordonnées du Client")
            # self.window.resizable(width=0,height=0)

            # entrer le nom
            self.nom = customtkinter.CTkEntry(master=self.windowC,

                                                placeholder_text="Nom")

            self.nom.grid(row=0, column=1, sticky="we", padx=30, pady=6)

            #DK#self.nom.entry.bind("<Return>", self.search_event)

            # entrer la longitude
            self.longi = customtkinter.CTkEntry(master=self.windowC,

                                                placeholder_text="Longitude")

            self.longi.grid(row=1, column=1, sticky="we", padx=30, pady=6)

            #DK#self.longi.entry.bind("<Return>", self.search_event)

            # entrer la latittude
            self.lati = customtkinter.CTkEntry(master=self.windowC,

                                                placeholder_text="Latittude")

            self.lati.grid(row=2, column=1, sticky="we", padx=30, pady=6)

            #DK#self.lati.entry.bind("<Return>", self.search_event)

            # entrer la Capacite du camion
            self.demande = customtkinter.CTkEntry(master=self.windowC,

                                                placeholder_text="Demande du client")

            self.demande.grid(row=3, column=1, sticky="we", padx=30, pady=6)

            #DK#self.demande.entry.bind("<Return>", self.set_marker_event)

            # Boutton valider

            self.valider_depot = customtkinter.CTkButton(master=self.windowC, text="Ajouter")

            self.valider_depot.configure(command=self.set_marker_eventC)

            self.valider_depot.grid(row=4, column=1, sticky="we", padx=30, pady=6)

#--------------------------------------------------------------------------------

#----------------------- AUTO CLIENT TOPLEVEL----------------------------

    def auto_client_toplevel(self,coordinates_tuple):
        
        if not self.sommet_list:
            # self.dialog = customtkinter.CTkInputDialog(master=None, text="Type in a number:")
            tkinter.messagebox.showerror(title='Erreur', message="Le depot n'est pas encore configuré; Configurer le depot")
            self.auto_depot_toplevel(coordinates_tuple)
        else:
            self.windowC = customtkinter.CTkToplevel(self)
            self.windowC.geometry("200x220")
            self.windowC.title("Coordonnées du Client")
            # self.window.resizable(width=0,height=0)

            # entrer la longitude
            self.long_label= customtkinter.CTkLabel(master= self.windowC, text="longitude: "+str(coordinates_tuple[0]))
            self.long_label.grid(row=0, column=1)
            self.longiti=coordinates_tuple[0]

            # entrer la latittude
            self.lat_label= customtkinter.CTkLabel(master= self.windowC, text="longitude: "+str(coordinates_tuple[1]))
            self.lat_label.grid(row=1, column=1)
            self.latiti=coordinates_tuple[1]

            # entrer le nom
            self.nom = customtkinter.CTkEntry(master=self.windowC,

                                                placeholder_text="Nom")

            self.nom.grid(row=2, column=1, sticky="we", padx=30, pady=6)

            #self.nom.entry.bind("<Return>", self.search_event)

            # entrer la Capacite du camion
            self.demande = customtkinter.CTkEntry(master=self.windowC,

                                                placeholder_text="Demande du client")

            self.demande.grid(row=3, column=1, sticky="we", padx=30, pady=6)

            #DK#self.demande.entry.bind("<Return>", self.auto_set_marker_eventC)

            # Boutton valider

            self.valider_depot = customtkinter.CTkButton(master=self.windowC, text="Ajouter")

            self.valider_depot.configure(command=self.auto_set_marker_eventC)

            self.valider_depot.grid(row=4, column=1, sticky="we", padx=30, pady=6)

#--------------------------------------------------------------------------------

#------------------------ TABLEAU CLIENTS----------------------------------------

    def tableau_clients(self):
        self.window_tab = customtkinter.CTkToplevel(self)
        self.window_tab.geometry(str(App.WIDTH - 200) + "x" + str(App.HEIGHT))
        self.window_tab.minsize(App.WIDTH - 200, App.HEIGHT)
        self.window_tab.title("Liste des clients")
        #self.window_tab.resizable(width=0, height=0)

        self.frame_top = customtkinter.CTkFrame(master=self.window_tab, width=150, corner_radius=0)

        self.frame_top.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")



        self.frame_bottom = customtkinter.CTkFrame(master=self.window_tab, corner_radius=0, fg_color=None)

        self.frame_bottom.grid(row=1, column=0, rowspan=1, pady=0, padx=0, sticky="nsew")


        self.frame_footer = customtkinter.CTkFrame(master=self.window_tab, corner_radius=0, fg_color=None)

        self.frame_footer.grid(row=2, column=0, rowspan=1, pady=0, padx=0, sticky="nsew")

        self.frame_coord = customtkinter.CTkFrame(master=self.window_tab, corner_radius=0, fg_color=None)

        self.frame_coord.grid(row=3, column=0, rowspan=1, pady=0, padx=0, sticky="nsew")

        

        def completer_tableau(self):

            self.tree= ttk.Treeview(self.frame_top, columns= (1,2,3), height= 18, show= 'headings')
            self.tree.grid(row=0,column=0, sticky='n')
            #self.tree.clear()

            self.tree.heading(1, text= 'Nom')
            self.tree.heading(2, text= 'Adresse')
            self.tree.heading(3, text= 'Quantité')

            for i in self.tree.get_children():
                self.tree.delete(i)
            for row in self.sommet_listC:
                #print(row)
                # lieu=tkintermapview.convert_coordinates_to_address(row[0], row[1])
                # lieu= str(lieu.street) + '/' + str(lieu.city) + '/' + str(lieu.country)
                lieu= self.adresse(row[0],row[1])
                rowz= [row[3],lieu, row[2]]
                self.tree.insert('', "end", values= rowz)
            #print("dk")

        completer_tableau(self)
        
        self.button_5 = customtkinter.CTkButton(master=self.frame_bottom,

                                                text="Importer une liste des clients",

                                                width=90,

                                                command=self.importer)

        self.button_5.grid(row=0, column=1, sticky="w", padx=(12, 0), pady=12)

        self.button_6 = customtkinter.CTkButton(master=self.frame_bottom,

                                                text="Enregistrer la liste des clients",

                                                width=90,

                                                command=self.creer_excel)

        self.button_6.grid(row=0, column=3, sticky="w", padx=(12, 0), pady=12)

        self.button_6 = customtkinter.CTkButton(master=self.frame_bottom,

                                                text="Actualiser",

                                                width=90,

                                                command=completer_tableau(self))

        self.button_6.grid(row=0, column=4, sticky="w", padx=(12, 0), pady=12)

        if self.sommet_list:
            self.depot_label = customtkinter.CTkLabel(self.frame_footer, 
                                                    text="Depot : " 
                                                    +self.adresse(self.sommet_list[0][0],self.sommet_list[0][1])
                                                    +"        Capacité du Camion : "
                                                    + str(self.sommet_list[0][2])
                                                    , anchor="n")

            self.depot_label.grid(row=0, column=0, padx=(20, 20), pady=(5, 0))
            self.coord_label = customtkinter.CTkLabel(self.frame_coord, 
                                                    text=""
                                                    +"Longitude= " +str(self.sommet_list[0][0]) +"  "
                                                    +"Lattitude= "+ str(self.sommet_list[0][1])
                                                    , anchor="n")

            self.coord_label.grid(row=0, column=0, padx=(20, 20), pady=(5, 0))

        

#--------------------------------- MATRICE DES DISTANCES -------------------------------------------------
    def matriceDeDistance(self):

        
        #print("Rute = ", rute.create_rute(0))
        def calcul_distance(a,b):

            eplb_park = a
            moulins_mairie = b
            # type de déplacement : cycle, foot, horse, tram, train, car ?
            router = Router("car")
            # Création des points de départ et d'arrivée
            depart = router.findNode(eplb_park[0],eplb_park[1])
            arrivee = router.findNode(moulins_mairie[0],moulins_mairie[1])
            # calcul itinéraire : test de l'existence d'une route
            status, itineraire = router.doRoute(depart, arrivee)
            if status == 'success':
                routeLatLons = list(map(router.nodeLatLon, itineraire)) # liste des points du parcours
                print(routeLatLons)

                longueur = 0
                i = 0
                while i < len (routeLatLons)-1:
                    module= sqrt((routeLatLons[i][0]-routeLatLons[i+1][0])*(routeLatLons[i][0]-routeLatLons[i+1][0])+(routeLatLons[i][1]-routeLatLons[i+1][1])*(routeLatLons[i][1]-routeLatLons[i+1][1]))
                    longueur += module
                    i = i+1
                return longueur

            else:
                x=a[0]-b[0]
                y=a[1]-b[1]
                ab=sqrt((x*x)+(y*y))
                return ab
        self.matrice_distance.clear()
        for i in self.sommet_list:
            #print('=======================')
            distance_ligne=[]
            distance_ligne.clear()
            for j in self.sommet_list:
                distance= calcul_distance(i,j)
                #print(distance)
                distance_ligne.append(distance)
                #print(distance_ligne)
            self.matrice_distance.append(distance_ligne)
            #print(self.matrice_distance)


        self.customer.clear()
        for i in self.sommet_listC:
            self.customer.append((i[3],i[2]))
        

        self.distance = self.matrice_distance
        # print(self.sommet_list)
        # print('===')
        #print(self.matrice_distance)
        # data = DataModel()
        # distance = Distance(data.customer_data_model(),
        #             data.create_distance_data_model())
        # rute = Rute(data.customer_data_model(), distance.execute_distance())
        # print("Rute = ", rute.create_rute(0))
        return self.matrice_distance

#==========================================================================================================

#--------------------------------- MATRICE DES ECONOMIES -------------------------------------------------

    def count_distance_customer_dinamic(self):
            counter=1
            self.matriceDeDistance()
            self.matrice_savings.clear()
            data_distance = self.matrice_distance
            # print('==============================')
            # print(data_distance)
            # print('==============================')
            leng_matrix = len(data_distance)
            count= counter+1

            while count < leng_matrix:
                counter_loop = counter + 1
                _data = []
                while counter_loop < leng_matrix:
                    #print(str(counter_loop) + '/' + str(leng_matrix))
                    x = data_distance[counter_loop][0]
                    #print(x)
                    y = data_distance[0][counter]
                    #print(y)
                    xy = data_distance[counter_loop][counter]
                    #print(xy)
                    total_xy = x + y - xy
                    #print(total_xy)
                    #_data.append(round(total_xy, 2))
                    #_data.append(total_xy)
                    self.matrice_savings.append(total_xy)

                    counter_loop += 1
                counter += 1
                #print(_data)
                count += 1
            #print(self.matrice_savings)
            return self.matrice_savings
#==========================================================================================================

#*********************************** VRP ******************************************************************

#           ++++++++++++++++++class Distance()++++++++++++++++++++++++
    def distances(self):
        self.matriceDeDistance()
        customer = []
        distance = []
        depot = 0

        customer = self.customer
        distance = self.distance
        # print(customer)
        # print(distance)

        def count_distance_customer_dinamic(counter):
            data_distance = distance
            leng_matrix = len(data_distance)
            counter_loop = counter + 1
            _data = []
            while counter_loop < leng_matrix:
                x = data_distance[counter_loop][0]
                y = data_distance[0][counter]
                xy = data_distance[counter_loop][counter]
                total_xy = x + y - xy
                _data.append(total_xy)

                counter_loop += 1
            return _data

        def execute_distance(self):
            starting_point = 1
            data_distance_matrix = []
            while starting_point < len(distance) - 1:
                data_distance_matrix.append(
                    count_distance_customer_dinamic(starting_point))
                starting_point += 1
            #print(data_distance_matrix)
            #print("=======================================")
            for i in data_distance_matrix:
                for j in range(len(customer) - len(i)):
                    i.insert(j, 0)
            print(data_distance_matrix)
            print("=======================================")
            return data_distance_matrix

        eco_matrix= execute_distance(self)
        return eco_matrix
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#                        ++++++++++class Rute+++++++++++++++++++
    def routes (self):
        customer = []
        distance = []
        rute = []

        
        customer = self.customer
        distance = self.distances()

        #print(distance)

        def create_node(self):

            temp_distance = []
            for i in distance:
                for j in i:
                    if j != 0:
                        temp_distance.append(j)
            temp_distance.sort(reverse=True)
            #print(temp_distance)
            counter = 0
            temp_node = []
            while counter < len(temp_distance):
                max_distance = temp_distance[counter]
                for index, k in enumerate(distance):
                    if max_distance in k:
                        node_one = (index + 1)
                        node_two = (k.index(max_distance) + 1)
                        if len(temp_node) == 0:
                            temp_node.extend([node_one, node_two])
                        else:

                            if node_one not in temp_node:
                                temp_node.append(node_one)

                            if node_two not in temp_node:
                                temp_node.append(node_two)

                            if node_one and node_two not in temp_node:
                                temp_node.extend([node_one, node_two])

                counter += 1
            #print('--------------------------------')
            #print(temp_node)
            return temp_node

        #create_node(self)

        def create_rute(self, index):
            node = create_node(self)
            max_capacity = self.sommet_list[0][2]
            print(max_capacity)
            order_capacity = 0
            temp_node = []

            for i in node[index:len(node)]:
                order_capacity += customer[i - 1][1]
                if order_capacity >= max_capacity:
                    rute.append(temp_node)
                    create_rute(self,node.index(i))
                    break
                else:
                    temp_node.append(i)

                if node.index(i) == (len(node) - 1):
                    rute.append(temp_node)
                    break

            return rute

        print('++++++++++++++++++++++')
        r= []
        r= create_rute(self,0)

        # Premier fomatage
        r2 = []
        mirror = []
        for elt in r :
            mirror.append(0)
            for subelt in elt :
               mirror.append(subelt)
            mirror.append(0)
            elt = mirror
            r2.append(mirror)
            mirror = []
        self.cluster = r2
        #Second formatage
        couples = []
        for elt in self.genetix() :
            length = len(elt)
            couple = []
            i = 0
            while (i < length-1) :
                couple.append(elt[i])
                couple.append(elt[i+1])
                couples.append(couple)
                couple  =[]
                i += 1

        print("Je print R"+str(r))
        print("Je print R2"+str(self.cluster))
        print("Je print couple"+str(couples))
        self.routeC= r
        self.route= couples

        def nearest_insert(self):
            return True

        def nearest_neighbor(self):
            return True

#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#============================== GENETIC ALGORITHM=========================================================

    def genetix(self):
        best_cluster= []
        for cluster in self.cluster:
            cluster.pop()
            if len(cluster) < 4 :
                #print(cluster)
                best_cluster.append(cluster)
            else:
                
                best_cluster.append(self.gen(cluster))
                #print(mat_distance)

        for l in best_cluster:
            l.append(0)
        #print(best_cluster)
        #print(self.cluster)

        return best_cluster
                




# ----------------------------- DESSINER ROUTE ------------------------------------------------------------

    def dessiner_route(self):

        self.clear_path()
        self.routes()
        couleur="#3317f4"
        c=0
        router = Router("foot")

        for i in self.route:
            x= i[0]
            y= i[1]
            
            if x==0:
                c+= 1
                couleur=self.random_couleur()

            
            point1= self.sommet_list[x]
            point2= self.sommet_list[y]

            depart = router.findNode(point1[0],point1[1])
            arrivee = router.findNode(point2[0],point2[1])
            # calcul itinéraire : test de l'existence d'une route
            status, itineraire = router.doRoute(depart, arrivee)
            if status == 'success':
                routeLatLons = list(map(router.nodeLatLon, itineraire)) # liste des points du parcours
                print(routeLatLons)
                 
                j=0
                while j < int(len(routeLatLons))-1:
        
                    self.polygone.append(self.map_widget.set_path([[routeLatLons[j][0],routeLatLons[j][1]],[routeLatLons[j+1][0],routeLatLons[j+1][1]]], width= 3, color=couleur))
                
                    j = j+1
            
            else:
                self.polygone.append(self.map_widget.set_path([[point1[0],point1[1]],[point2[0],point2[1]]], width= 3, color=couleur))

#----------------------------------------------------------------------------------------------------------

#---------------------------- ADRESSE --------------------------------------------------------------------

    def adresse(self,x,y):
        long = x
        lat = y
        add = tkintermapview.convert_coordinates_to_address(long,lat)
        num = add.housenumber
        av = add.street
        ville = add.city
        pays = add.country
        fasi = ""

        def adresse_cheker(addr):
            if addr == "None":
                return ""
            else: 
                return addr

        if pays :
            if ville: 
                fasi = ville +""+ fasi
                if av:  
                    fasi = av +" "+ fasi
                    if num:
                        fasi = num +" "+ fasi
                        return fasi
            else: 
                fasi = pays +" "+ fasi
                if av:  
                    fasi = av +" "+ fasi
                    if num:
                        fasi = num +" "+ fasi
                        return fasi

        return fasi
        
        

#---------------------------------------------------------------------------------------------------------

# ------------------------------  IMPORTER --------------------------------------------------------------

    def importer(self):
        if self.sommet_list:

            filetypes= (("tableau", "*.xlsx"),("vrp project","*.vrpp"))
            #file = filedialog.askopenfile(master= self.window_tab, initialdir="/", title= "select", filetypes= filetypes)
            file_name = filedialog.askopenfilename(master= self.window_tab, initialdir="/", title= "select", filetypes= filetypes)
            #file_name = filedialog.LoadFileDialog(master= self.window_tab, initialdir="/", title= "select", filetypes= filetypes)
            sheet= pandas.read_excel(file_name)
            print(file_name)

            self.wb = load_workbook(file_name)
            ws = self.wb["Liste des Clients"]
            x=1
            y=0

            for c in self.sommet_list:
                if self.sommet_list.index(c) != 0:
                    del c
            self.sommet_listC.clear()

            for row in ws.values:
                client= []
                for cellule in row:
                    #if int(cellule[2]) < int(self.sommet_list[0][2]): #Tester si la valeur demandee ne depasse pas la capaticite du camion
                        client.append(cellule)

                self.sommet_list.append(client)
                self.sommet_listC.append(client)

            while y < len(self.sommet_listC):

                self.marker_list.append(self.map_widget.set_marker(self.sommet_listC[y][0], 
                                        self.sommet_listC[y][1], text=self.sommet_listC[y][3],
                                        marker_color_outside= "#cc19cc",marker_color_circle="#eef",
                                        text_color="#A2A"))
                y = y+1

        else:
            tkinter.messagebox.showerror(title='Erreur', message="Le depot n'est pas encore configuré; Configurer le depot")

    def creer_excel(self):
            ws = self.wb.create_sheet("Liste des Clients", 0)
            for x in range(1,len(self.sommet_list)):
                for y in range(1,len(self.sommet_list[x])+1):
                    ws.cell(row=x, column=y, value= self.sommet_list[x][y-1]) 
            for row in ws.values:    
                for value in row:      
                    print(value)
            
            filetypes= (("tableau", "*.xlsx"),("vrp project","*.vrpp"))
            #file_name = filedialog.askopenfilename(master= self.window_tab, initialdir="/", title= "select", filetypes= filetypes)
            file_name = filedialog.asksaveasfilename(title= 'Exporter la liste des clients', defaultextension='*.xlsx', filetypes=filetypes,confirmoverwrite=1)

            self.wb.save(file_name)
#-------------------------------------------------------------------------------------------------------

    def random_color(self):
        #couleur=""
        r=1
        r= random.randint(1,6)
        #print(r)

        if r==1:
            return "#CCC"
        if r==2:
            return "#00f"
        if r==3:
            return "#0f0"
        if r==4:
            return "#f00"
        if r==5:
            return "#ff0"
        if r==6:
            return "#f0f"
        if r==r:
            return "#0ff"
    
    def random_couleur(self):
        couleur="#"
        r= random.randint(1048576,16777215)
        h= hex(r)[2::]
        c= list(h)
        #print(c)
        c.insert(0,"#")
        couleur= "".join(c)
        #print(couleur)
        return couleur

    def position_on_depot(self):
        if self.sommet_list:
            if self.sommet_list[0]:
                #position=(self.sommet_list[0][0],self.sommet_list[0][1])
                self.map_widget.set_position(self.sommet_list[0][0],self.sommet_list[0][1])

#**********************************************************************************************************

    def gen(self, cluster):

        liste_client = cluster
        print (liste_client)


        class City:
            def __init__(self, x=None, y=None):
                self.x = None
                self.y = None
                if x is not None:
                    self.x = x
                else:
                    self.x = int(random.random() * 200)
                if y is not None:
                    self.y = y
                else:
                    self.y = int(random.random() * 200)
            
            def getX(self):
                return self.x
            
            def getY(self):
                return self.y
            
            def distanceTo(self, city):
                xDistance = abs(self.getX() - city.getX())
                yDistance = abs(self.getY() - city.getY())
                distance = sqrt( (xDistance*xDistance) + (yDistance*yDistance) )
                return distance
            
            def __repr__(self):
                return str(self.getX()) + ", " + str(self.getY())


        class TourManager:
            destinationCities = []
            
            def addCity(self, city):
                self.destinationCities.append(city)
            
            def getCity(self, index):
                return self.destinationCities[index]
            
            def numberOfCities(self):
                return len(self.destinationCities)


        class Tour:
            def __init__(self, tourmanager, tour=None):
                self.tourmanager = tourmanager
                self.tour = []
                self.fitness = 0.0
                self.distance = 0
                if tour is not None:
                    self.tour = tour
                else:
                    for i in range(0, self.tourmanager.numberOfCities()):
                        self.tour.append(None)
            
            def __len__(self):
                return len(self.tour)
            
            def __getitem__(self, index):
                return self.tour[index]
            
            def __setitem__(self, key, value):
                self.tour[key] = value
            
            def __repr__(self):
                geneString = "|"
                for i in range(0, self.tourSize()):
                    geneString += str(self.getCity(i)) + "|"
                return geneString
            
            def generateIndividual(self):
                for cityIndex in range(0, self.tourmanager.numberOfCities()):
                    self.setCity(cityIndex, self.tourmanager.getCity(cityIndex))
                random.shuffle(self.tour)
            
            def getCity(self, tourPosition):
                return self.tour[tourPosition]
            
            def setCity(self, tourPosition, city):
                self.tour[tourPosition] = city
                self.fitness = 0.0
                self.distance = 0
            
            def getFitness(self):
                if self.fitness == 0:
                    self.fitness = 1/float(self.getDistance())
                return self.fitness
            
            def getDistance(self):
                if self.distance == 0:
                    tourDistance = 0
                    for cityIndex in range(0, self.tourSize()):
                        fromCity = self.getCity(cityIndex)
                        destinationCity = None
                        if cityIndex+1 < self.tourSize():
                            destinationCity = self.getCity(cityIndex+1)
                        else:
                            destinationCity = self.getCity(0)
                            tourDistance += fromCity.distanceTo(destinationCity)
                    self.distance = tourDistance
                return self.distance
            
            def tourSize(self):
                return len(self.tour)
            
            def containsCity(self, city):
                return city in self.tour


        class Population:
            def __init__(self, tourmanager, populationSize, initialise):
                self.tours = []
                for i in range(0, populationSize):
                    self.tours.append(None)
                
                if initialise:
                    for i in range(0, populationSize):
                        newTour = Tour(tourmanager)
                        newTour.generateIndividual()
                        self.saveTour(i, newTour)
                
            def __setitem__(self, key, value):
                self.tours[key] = value
            
            def __getitem__(self, index):
                return self.tours[index]
            
            def saveTour(self, index, tour):
                self.tours[index] = tour
            
            def getTour(self, index):
                return self.tours[index]
            
            def getFittest(self):
                fittest = self.tours[0]
                for i in range(0, self.populationSize()):
                    if fittest.getFitness() <= self.getTour(i).getFitness():
                        fittest = self.getTour(i)
                return fittest
            
            def populationSize(self):
                return len(self.tours)


        class GA:
            def __init__(self, tourmanager):
                self.tourmanager = tourmanager
                self.mutationRate = 0.015
                self.tournamentSize = 5
                self.elitism = True
            
            def evolvePopulation(self, pop):
                newPopulation = Population(self.tourmanager, pop.populationSize(), False)
                elitismOffset = 0
                if self.elitism:
                    newPopulation.saveTour(0, pop.getFittest())
                    elitismOffset = 1
                
                for i in range(elitismOffset, newPopulation.populationSize()):
                    parent1 = self.tournamentSelection(pop)
                    parent2 = self.tournamentSelection(pop)
                    child = self.crossover(parent1, parent2)
                    newPopulation.saveTour(i, child)
                
                for i in range(elitismOffset, newPopulation.populationSize()):
                    self.mutate(newPopulation.getTour(i))
                
                return newPopulation
            
            def crossover(self, parent1, parent2):
                child = Tour(self.tourmanager)
                
                startPos = int(random.random() * parent1.tourSize())
                endPos = int(random.random() * parent1.tourSize())
                
                for i in range(0, child.tourSize()):
                    if startPos < endPos and i > startPos and i < endPos:
                        child.setCity(i, parent1.getCity(i))
                    elif startPos > endPos:
                        if not (i < startPos and i > endPos):
                            child.setCity(i, parent1.getCity(i))
                
                for i in range(0, parent2.tourSize()):
                    if not child.containsCity(parent2.getCity(i)):
                        for ii in range(0, child.tourSize()):
                            if child.getCity(ii) == None:
                                child.setCity(ii, parent2.getCity(i))
                                break
                
                return child
            
            def mutate(self, tour):
                for tourPos1 in range(0, tour.tourSize()):
                    if random.random() < self.mutationRate:
                        tourPos2 = int(tour.tourSize() * random.random())
                        
                        city1 = tour.getCity(tourPos1)
                        city2 = tour.getCity(tourPos2)
                        
                        tour.setCity(tourPos2, city1)
                        tour.setCity(tourPos1, city2)
            
            def tournamentSelection(self, pop):
                tournament = Population(self.tourmanager, self.tournamentSize, False)
                for i in range(0, self.tournamentSize):
                    randomId = int(random.random() * pop.populationSize())
                    tournament.saveTour(i, pop.getTour(randomId))
                fittest = tournament.getFittest()
                return fittest



        
            
        tourmanager = TourManager()
            
        # Create and add our cities
        for vil in liste_client:
            print(vil)
            city = City(self.sommet_list[vil][0], self.sommet_list[vil][1])
            tourmanager.addCity(city)
        
        print(liste_client)
        
            # Initialize population
        pop = Population(tourmanager, 50, True)
        initial_distance= pop.getFittest().getDistance()
        print ("Initial distance: " + str(pop.getFittest().getDistance()))
            
        # Evolve population for 50 generations
        ga = GA(tourmanager)
        pop = ga.evolvePopulation(pop)
        for i in range(0, 100):
            pop = ga.evolvePopulation(pop)
            
        # Print final results
        print ("Finished")
        final_distance= pop.getFittest().getDistance()
        print ("Final distance: " + str(pop.getFittest().getDistance()))
        print ("Solution:")
        new_cluster_coord = pop.getFittest()
        print (pop.getFittest())

        new_cluster = []

        if final_distance < initial_distance:

            for ville in new_cluster_coord:
                for place in self.sommet_list:
                    if ville.getX() in place:
                        indice = self.sommet_list.index(place)
                        if indice != 0:
                            new_cluster.append(indice)
                
            
            print(new_cluster)
            new_cluster.insert(0,0)
            #best_cluster.append(new_cluster)
            
            return new_cluster

        else:

            return liste_client


    def on_closing(self, event=0):

        self.destroy()



    def start(self):

        self.mainloop()



###########################


if __name__ == "__main__":

    app = App()

    app.start()